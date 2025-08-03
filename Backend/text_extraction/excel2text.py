import openpyxl
import os

def extract_text_from_excel(excel_path):
    if not os.path.exists(excel_path):
        return "Error: File not found."

    try:
        workbook = openpyxl.load_workbook(excel_path)
        extracted_text = []
        for sheet in workbook.sheetnames:
            extracted_text.append(f"Sheet: {sheet}")
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) if cell else "" for cell in row])
                extracted_text.append(row_text)
        return "\n".join(extracted_text)
    except Exception as e:
        return f"Error reading the Excel file - {e}"

if __name__ == "__main__":
    excel_path = input("Enter the path to the Excel file (.xlsx): ").strip()
    extract_text_from_excel(excel_path)
