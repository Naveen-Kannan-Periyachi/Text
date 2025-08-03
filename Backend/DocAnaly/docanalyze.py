import os
import fitz  # PyMuPDF
import nltk
import easyocr
from docx import Document
import openpyxl
from PIL import Image
from nltk import sent_tokenize, word_tokenize, pos_tag
from nltk.corpus import stopwords
from nltk.sentiment import SentimentIntensityAnalyzer
import spacy

# Download required NLTK data (only once needed)
nltk.download('punkt')
nltk.download('averaged_perceptron_tagger')
nltk.download('stopwords')
nltk.download('vader_lexicon')

# ---------- Text Extraction Functions ----------
def extract_from_txt(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        return f.read()

def extract_from_docx(file_path):
    doc = Document(file_path)
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())

def extract_from_pdf(file_path):
    doc = fitz.open(file_path)
    return "\n".join([page.get_text() for page in doc])

def extract_from_image(file_path):
    reader = easyocr.Reader(['en'])
    results = reader.readtext(file_path)
    return "\n".join([text for _, text, _ in results])

def extract_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    text = []
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(values_only=True):
            line = " | ".join([str(cell) if cell else '' for cell in row])
            text.append(line)
    return "\n".join(text)

def detect_and_extract(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.txt':
        return extract_from_txt(file_path)
    elif ext == '.docx':
        return extract_from_docx(file_path)
    elif ext == '.pdf':
        return extract_from_pdf(file_path)
    elif ext in ['.jpg', '.jpeg', '.png']:
        return extract_from_image(file_path)
    elif ext in ['.xlsx']:
        return extract_from_excel(file_path)
    else:
        raise ValueError("Unsupported file type: " + ext)

def analyze_text(text):
    result = []

    # Sentence tokenization
    sentences = sent_tokenize(text)
    result.append(f"📌 Sentence Tokenization:\nTotal sentences: {len(sentences)}\n")

    # Word tokenization
    words = word_tokenize(text)
    result.append(f"📌 Word Tokenization:\nTotal words: {len(words)}\n")

    # POS tagging
    tagged = pos_tag(words)
    result.append("📌 POS Tagging (first 20):\n" + str(tagged[:20]) + "\n")

    # Stopword removal
    stop_words = set(stopwords.words("english"))
    filtered_words = [w for w in words if w.lower() not in stop_words and w.isalpha()]
    result.append(f"\n📌 Filtered Words (no stopwords, first 20):\n{filtered_words[:20]}\n")

    # Sentiment analysis
    sia = SentimentIntensityAnalyzer()
    sentiment = sia.polarity_scores(text)
    result.append("📌 Sentiment Analysis:\n" + str(sentiment))

    # Named Entity Recognition (NER)
    try:
        nlp = spacy.load("en_core_web_sm")
        doc = nlp(text)
        entities = [(ent.text, ent.label_) for ent in doc.ents]
        if entities:
            result.append("📌 Named Entities (first 20):\n" + str(entities[:20]))
        else:
            result.append("📌 Named Entities: None found.")
    except Exception as e:
        result.append(f"📌 Named Entity Recognition: FAILED - {e}")

    return "\n".join(result)

# ---------- Main ----------
def main():
    file_path = input("Enter the path to the file: ").strip()
    if not os.path.exists(file_path):
        print("❌ File not found.")
        return

    print("\n🔍 Extracting text from file...")
    try:
        text = detect_and_extract(file_path)
    except Exception as e:
        print("❌ Error extracting text:", e)
        return

    if not text.strip():
        print("❌ No text found.")
        return

    analysis = analyze_text(text)

    # Print to terminal
    print("\n📊 Document Analysis Results:\n")
    print(analysis)

    # Save to file
    with open("document_analysis.txt", "w", encoding="utf-8") as f:
        f.write(analysis)

    print("\n✅ Document analysis complete. Results saved to 'document_analysis.txt'.")

if __name__ == "__main__":
    main()
