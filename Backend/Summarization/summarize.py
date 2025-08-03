import os
import mimetypes

# Text extraction libraries
from docx import Document
import fitz  # PyMuPDF
import easyocr
import openpyxl
from PIL import Image

# Summarization libraries
from transformers import pipeline
from sumy.parsers.plaintext import PlaintextParser
from sumy.nlp.tokenizers import Tokenizer
from sumy.summarizers.lex_rank import LexRankSummarizer
import torch

def extract_from_txt(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        return {1: f.read()}

def extract_from_docx(file_path):
    doc = Document(file_path)
    text = "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
    return {1: text}

def extract_from_pdf(file_path):
    doc = fitz.open(file_path)
    return {i+1: page.get_text() for i, page in enumerate(doc)}

def extract_from_image(file_path):
    reader = easyocr.Reader(['en'], gpu=True if os.environ.get("USE_CUDA", "1") == "1" else False)
    results = reader.readtext(file_path)
    text = "\n".join([text for _, text, _ in results])
    return {1: text}

def extract_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    text = []
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(values_only=True):
            line = " | ".join([str(cell) if cell else '' for cell in row])
            text.append(line)
    return {1: "\n".join(text)}

def detect_and_extract(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.txt':
        return extract_from_txt(file_path)
    elif ext == '.docx':
        return extract_from_docx(file_path)
    elif ext == '.pdf':
        return extract_from_pdf(file_path)
    elif ext in ['.jpg', '.jpeg', '.png']:
        return extract_from_image(file_path)
    elif ext in ['.xlsx']:
        return extract_from_excel(file_path)
    else:
        raise ValueError("Unsupported file type: " + ext)

def summarize_abstractive(text):
    # Defensive: Truncate to 4000 chars (model limit)
    if not text.strip():
        return "No text to summarize."
    max_input_length = 4000  # chars, adjust as needed
    if len(text) > max_input_length:
        text = text[:max_input_length]
    summarizer = pipeline("summarization", model="facebook/bart-large-cnn", device=0 if torch.cuda.is_available() else -1)
    return summarizer(text, max_length=130, min_length=30, do_sample=False)[0]['summary_text']

def summarize_extractive(text, sentence_count=3):
    parser = PlaintextParser.from_string(text, Tokenizer("english"))
    summarizer = LexRankSummarizer()
    summary = summarizer(parser.document, sentence_count)
    return " ".join([str(sentence) for sentence in summary])

def main():
    file_path = input("Enter the path to the file: ").strip()
    if not os.path.exists(file_path):
        print("File not found.")
        return

    print("\n🔍 Extracting text from:", file_path)
    try:
        extracted_text_by_page = detect_and_extract(file_path)
    except Exception as e:
        print("Error extracting text:", e)
        return

    if not any(text.strip() for text in extracted_text_by_page.values()):
        print("No text found in the file.")
        return

    print("\n📝 Extracted Text Preview:")
    for page, text in extracted_text_by_page.items():
        print(f"Page {page} Preview:\n{text[:500]}\n{'...' if len(text) > 500 else ''}\n")

    extractive_summaries = {}
    abstractive_summaries = {}

    for page, text in extracted_text_by_page.items():
        if not text.strip():
            continue
        print(f"\n📌 Performing Extractive Summarization for Page {page}...")
        try:
            extractive_summaries[page] = summarize_extractive(text)
        except Exception as e:
            extractive_summaries[page] = f"FAILED: {e}"

    for page, text in extracted_text_by_page.items():
        if not text.strip():
            continue
        print(f"\n📌 Performing Abstractive Summarization for Page {page}...")
        try:
            abstractive_summaries[page] = summarize_abstractive(text)
        except Exception as e:
            abstractive_summaries[page] = f"FAILED: {e}"

    # Save extractive summary
    with open("extractive_summary.txt", "w", encoding="utf-8") as f:
        for page in sorted(extractive_summaries):
            f.write(f"Page {page}:\n{extractive_summaries[page]}\n\n")

    # Save abstractive summary
    with open("abstractive_summary.txt", "w", encoding="utf-8") as f:
        for page in sorted(abstractive_summaries):
            f.write(f"Page {page}:\n{abstractive_summaries[page]}\n\n")

    print("\n✅ Summaries saved to 'extractive_summary.txt' and 'abstractive_summary.txt'.")

if __name__ == "__main__":
    main()
