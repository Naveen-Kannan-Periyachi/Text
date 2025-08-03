import os
import mimetypes
from docx import Document
import fitz  # PyMuPDF
import easyocr
import openpyxl
from deep_translator import GoogleTranslator
from PIL import Image

# Language name mapping for user display
LANGUAGES = {
    'hi': 'hindi', 'ta': 'tamil', 'fr': 'french', 'de': 'german', 'es': 'spanish', 'zh-CN': 'chinese (simplified)',
    'ar': 'arabic', 'ru': 'russian', 'en': 'english', 'it': 'italian', 'ja': 'japanese', 'ko': 'korean'
}

def extract_from_txt(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        return f.read()

def extract_from_docx(file_path):
    doc = Document(file_path)
    return "\n".join([p.text for p in doc.paragraphs if p.text.strip()])

def extract_from_pdf(file_path):
    doc = fitz.open(file_path)
    return "\n".join([page.get_text() for page in doc])

def extract_from_image(file_path):
    reader = easyocr.Reader(['en'])
    results = reader.readtext(file_path)
    return "\n".join([text for _, text, _ in results])

def extract_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    text = []
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(values_only=True):
            line = " | ".join([str(cell) if cell else '' for cell in row])
            text.append(line)
    return "\n".join(text)

def detect_and_extract(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.txt':
        return extract_from_txt(file_path)
    elif ext == '.docx':
        return extract_from_docx(file_path)
    elif ext == '.pdf':
        return extract_from_pdf(file_path)
    elif ext in ['.jpg', '.jpeg', '.png']:
        return extract_from_image(file_path)
    elif ext in ['.xlsx']:
        return extract_from_excel(file_path)
    else:
        raise ValueError("Unsupported file type: " + ext)

def translate_text(text, languages):
    translations = {}
    chunk_size = 4500  # Use a safe chunk size to avoid API limits

    def split_into_chunks(text, size):
        lines = text.splitlines()
        chunks, current_chunk = [], ""
        for line in lines:
            if len(current_chunk) + len(line) + 1 < size:
                current_chunk += line + "\n"
            else:
                chunks.append(current_chunk.strip())
                current_chunk = line + "\n"
        if current_chunk:
            chunks.append(current_chunk.strip())
        return chunks

    chunks = split_into_chunks(text, chunk_size)

    for lang in languages:
        translated_chunks = []
        try:
            for chunk in chunks:
                translated = GoogleTranslator(source='auto', target=lang).translate(chunk)
                translated_chunks.append(translated)
            translations[lang] = "\n".join(translated_chunks)
        except Exception as e:
            translations[lang] = f"Error translating to {lang}: {e}"

    return translations

def translate_file_api(file_path, language_codes):
    """
    API-ready function to extract text from file and translate to given languages.
    Returns a dict: {lang_code: translation or error}
    """
    if not os.path.exists(file_path):
        return {"error": "File not found."}
    try:
        extracted_text = detect_and_extract(file_path)
    except Exception as e:
        return {"error": f"Error extracting text: {e}"}
    if not extracted_text.strip():
        return {"error": "No text found in the file."}
    # Validate language codes
    languages = [lang for lang in language_codes if lang in LANGUAGES]
    if not languages:
        return {"error": "No valid languages entered."}
    translations = translate_text(extracted_text, languages)
    return translations

def main():
    file_path = input("Enter the path to the file: ").strip()
    if not os.path.exists(file_path):
        print("File not found.")
        return

    print("\n🔍 Extracting text from:", file_path)
    try:
        extracted_text = detect_and_extract(file_path)
    except Exception as e:
        print("Error extracting text:", e)
        return

    if not extracted_text.strip():
        print("No text found in the file.")
        return

    print("\n📝 Extracted Text Preview:\n" + extracted_text[:1000] + ("..." if len(extracted_text) > 1000 else ""))

    print("\n🌐 Supported languages:")
    for code, name in LANGUAGES.items():
        print(f"{code}: {name}")

    user_langs = input("\nEnter the language codes to translate to (comma-separated): ").strip()
    languages = [lang.strip() for lang in user_langs.split(',') if lang.strip() in LANGUAGES]

    if not languages:
        print("No valid languages entered.")
        return

    print("\n🌐 Translating text into:", ", ".join(languages))

    translations = translate_text(extracted_text, languages)

    with open("translated_output.txt", "w", encoding="utf-8") as f:
        for lang, translated in translations.items():
            f.write(translated + "\n\n")

    print("\n✅ Translations saved to 'translated_output.txt'.")

if __name__ == "__main__":
    main()
